import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

DB_CONFIG = {
    'user': 'root',
    'password': 'root',
    'host': 'localhost',
    'database': 'delta',
    'port': '3306'
}

EXCEL_CONFIG = {
    'file_path': 'D:/PRojetos/Python/automacao_planilha/ganhe_jogando.xlsx',
    'sheet_name': 'form',
    'id': 'id',
    'new_columns': ['minutesOnline', 'rank'] 
}

def conectar_db():
    """Conexão com MariaDB"""
    conn_str = f"mysql+pymysql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    return create_engine(conn_str)

def atualizar_excel():
    try:
        # 1. Ler dados existentes do Excel
        df_planilha = pd.read_excel(
            EXCEL_CONFIG['file_path'],
            sheet_name=EXCEL_CONFIG['sheet_name']
        )

        # 2. Consultar banco de dados
        engine = conectar_db()
        ids = df_planilha[EXCEL_CONFIG['id']].tolist()
        
        query = f"""
            SELECT {EXCEL_CONFIG['id']}, `minutesOnline`
            FROM vrp_users 
            WHERE {EXCEL_CONFIG['id']} IN ({','.join(map(str, ids))})
        """
        
        df_banco = pd.read_sql(query, engine)
        df_banco['rank'] = df_banco['minutesOnline'].rank(ascending=False, method='dense')

        # 3. Mesclar dados
        df_final = df_planilha.merge(
            df_banco,
            on=EXCEL_CONFIG['id'],
            how='left'
        )

        # 4. Carregar workbook existente
        book = load_workbook(EXCEL_CONFIG['file_path'])
        sheet = book[EXCEL_CONFIG['sheet_name']]

        # 5. Encontrar posições das colunas
        header = [cell.value for cell in sheet[1]]
        
        # Adicionar novas colunas se necessário
        for col in EXCEL_CONFIG['new_columns']:
            if col not in header:
                sheet.cell(row=1, column=len(header)+1, value=col)
                header.append(col)

        # Mapear colunas
        col_positions = {col: header.index(col)+1 for col in EXCEL_CONFIG['new_columns'] + [EXCEL_CONFIG['id']]}

        # 6. Atualizar dados mantendo formatação
        for idx, row in df_final.iterrows():
            for col in EXCEL_CONFIG['new_columns']:
                cell = sheet.cell(
                    row=idx+2,  # Assumindo que a linha 1 é o cabeçalho
                    column=col_positions[col],
                    value=row[col]
                )

        # 7. Salvar alterações
        book.save(EXCEL_CONFIG['file_path'])
        print("Excel atualizado com sucesso mantendo a formatação!")

    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    atualizar_excel()